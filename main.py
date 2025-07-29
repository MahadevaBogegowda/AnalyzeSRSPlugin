from fastapi import FastAPI, File, UploadFile
import openai
import tempfile
import os
import docx2txt
import fitz  # PyMuPDF
from openpyxl import load_workbook

app = FastAPI()

openai.api_key = os.getenv("OPENAI_API_KEY")

def extract_text_from_file(file_path, extension):
    if extension == "docx":
        return docx2txt.process(file_path)
    elif extension == "pdf":
        text = ""
        pdf = fitz.open(file_path)
        for page in pdf:
            text += page.get_text()
        return text
    elif extension == "txt":
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()
    elif extension == "xlsx":
        wb = load_workbook(file_path)
        text = ""
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                text += row_text + "\n"
        return text
    else:
        return None

@app.post("/analyze-srs")
async def analyze_srs(file: UploadFile = File(...)):
    file_ext = file.filename.lower().split('.')[-1]
    with tempfile.NamedTemporaryFile(delete=False, suffix=f".{file_ext}") as tmp:
        tmp.write(await file.read())
        tmp_path = tmp.name
    try:
        text = extract_text_from_file(tmp_path, file_ext)
        if not text:
            return {"error": f"Unsupported or empty file format: {file_ext}"}

        text = text[:3000]

        prompt = (
            "You are a business analyst. Convert the following software requirements "
            "into Agile user stories using this format:\n"
            "As a <user>, I want to <action>, so that <goal>.\n\n"
            f"Requirements:\n{text}"
        )

        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=[{"role": "user", "content": prompt}],
            temperature=0.3,
            max_tokens=600
        )

        return {"userStories": response.choices[0].message.content.strip()}
    
    finally:
        os.remove(tmp_path)
